import re
import os
import json
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import argparse
from collections import defaultdict
from scipy.stats import linregress
import matplotlib.ticker as mtick
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def read_file(file_path):
    """Lit le fichier texte extrait du PDF."""
    print(f"📂 Lecture du fichier : {file_path}")
    with open(file_path, 'r', encoding='utf-8') as file:
        text = file.read()
    return text

def clean_text(text):
    """Nettoie le texte en supprimant les caractères spéciaux et les sections inutiles."""
    print("🧹 Nettoyage du texte...")
    
    # Suppression des en-têtes et pieds de page répétitifs
    text = re.sub(r'Tesla, Inc. \/ Form 10-K \/ December 31, 2024.*?http:\/\/www\.sec\.gov', '', text, flags=re.DOTALL)
    
    # Suppression des numéros de page
    text = re.sub(r'\n\s*\d+\s*\n', '\n', text)
    
    return text

def extract_key_metrics(text):
    """Extrait les métriques financières clés du texte."""
    print("📊 Extraction des métriques financières clés...")
    
    metrics = {
        'revenue': {},
        'net_income': {},
        'gross_profit': {},
        'total_assets': {},
        'total_liabilities': {},
        'operating_cash_flow': {}
    }
    
    # Extraction des revenus
    # Pattern pour Tesla
    revenue_pattern = r"Total revenues\s*\$?\s*([\d,]+)\s*\$?\s*([\d,]+)\s*\$?\s*([\d,]+)"
    revenue_match = re.search(revenue_pattern, text)
    
    # Pattern pour Apple
    apple_revenue_pattern = r"Total net sales\s*\$?([\d,]+)\s*\$?([\d,]+)\s*\$?([\d,]+)"
    apple_revenue_match = re.search(apple_revenue_pattern, text)
    
    # Pattern pour Microsoft
    microsoft_revenue_pattern = r"Total revenue\s*\$?([\d,]+)\s*\$?([\d,]+)\s*\$?([\d,]+)"
    microsoft_revenue_match = re.search(microsoft_revenue_pattern, text)
    
    if revenue_match:
        metrics['revenue']['2024'] = revenue_match.group(1)
        metrics['revenue']['2023'] = revenue_match.group(2)
        metrics['revenue']['2022'] = revenue_match.group(3)
        print(f"✅ Revenus totaux extraits avec succès : {metrics['revenue']}")
    elif apple_revenue_match:
        metrics['revenue']['2024'] = apple_revenue_match.group(1)
        metrics['revenue']['2023'] = apple_revenue_match.group(2)
        metrics['revenue']['2022'] = apple_revenue_match.group(3)
        print(f"✅ Revenus totaux extraits avec succès (Apple) : {metrics['revenue']}")
    elif microsoft_revenue_match:
        metrics['revenue']['2024'] = microsoft_revenue_match.group(1)
        metrics['revenue']['2023'] = microsoft_revenue_match.group(2)
        metrics['revenue']['2022'] = microsoft_revenue_match.group(3)
        print(f"✅ Revenus totaux extraits avec succès (Microsoft) : {metrics['revenue']}")
    else:
        print("❌ Impossible de trouver les revenus totaux")
    
    # Extraction du bénéfice net
    # Pattern pour Tesla
    net_income_pattern = r"Net income\s*\$?\s*([\d,]+)\s*\$?\s*([\d,]+)\s*\$?\s*([\d,]+)"
    net_income_match = re.search(net_income_pattern, text)
    
    # Pattern alternatif pour Apple et Microsoft
    alt_net_income_pattern = r"Net income.*?(\$[\d,]+)\s+(\$[\d,]+)\s+(\$[\d,]+)"
    alt_net_income_match = re.search(alt_net_income_pattern, text)
    
    if net_income_match:
        metrics['net_income']['2024'] = net_income_match.group(1)
        metrics['net_income']['2023'] = net_income_match.group(2)
        metrics['net_income']['2022'] = net_income_match.group(3)
        print(f"✅ Bénéfice net extrait avec succès : {metrics['net_income']}")
    elif alt_net_income_match:
        metrics['net_income']['2024'] = alt_net_income_match.group(1)
        metrics['net_income']['2023'] = alt_net_income_match.group(2)
        metrics['net_income']['2022'] = alt_net_income_match.group(3)
        print(f"✅ Bénéfice net extrait avec succès : {metrics['net_income']}")
    else:
        print("❌ Impossible de trouver le bénéfice net")
    
    # Extraction du bénéfice brut
    # Pattern pour Tesla
    gross_profit_pattern = r"Gross profit\s*\$?\s*([\d,]+)\s*\$?\s*([\d,]+)\s*\$?\s*([\d,]+)"
    gross_profit_match = re.search(gross_profit_pattern, text)
    
    # Pattern alternatif pour Apple et Microsoft
    alt_gross_profit_pattern = r"Gross margin\s*\$?([\d,]+)\s*\$?([\d,]+)\s*\$?([\d,]+)"
    alt_gross_profit_match = re.search(alt_gross_profit_pattern, text)
    
    if gross_profit_match:
        metrics['gross_profit']['2024'] = gross_profit_match.group(1)
        metrics['gross_profit']['2023'] = gross_profit_match.group(2)
        metrics['gross_profit']['2022'] = gross_profit_match.group(3)
        print(f"✅ Bénéfice brut extrait avec succès : {metrics['gross_profit']}")
    elif alt_gross_profit_match:
        metrics['gross_profit']['2024'] = alt_gross_profit_match.group(1)
        metrics['gross_profit']['2023'] = alt_gross_profit_match.group(2)
        metrics['gross_profit']['2022'] = alt_gross_profit_match.group(3)
        print(f"✅ Bénéfice brut extrait avec succès : {metrics['gross_profit']}")
    else:
        print("❌ Impossible de trouver le bénéfice brut")
    
    # Extraction des actifs totaux
    # Pattern pour Tesla
    total_assets_pattern = r"Total assets\s*\$?\s*([\d,]+)\s*\$?\s*([\d,]+)"
    total_assets_match = re.search(total_assets_pattern, text)
    
    # Pattern alternatif pour Apple et Microsoft
    alt_total_assets_pattern = r"Total assets.*?(\$[\d,]+)\s+(\$[\d,]+)"
    alt_total_assets_match = re.search(alt_total_assets_pattern, text)
    
    if total_assets_match:
        metrics['total_assets']['2024'] = total_assets_match.group(1)
        metrics['total_assets']['2023'] = total_assets_match.group(2)
        print(f"✅ Actifs totaux extraits avec succès : {metrics['total_assets']}")
    elif alt_total_assets_match:
        metrics['total_assets']['2024'] = alt_total_assets_match.group(1)
        metrics['total_assets']['2023'] = alt_total_assets_match.group(2)
        print(f"✅ Actifs totaux extraits avec succès : {metrics['total_assets']}")
    else:
        print("❌ Impossible de trouver les actifs totaux")
    
    # Extraction des passifs totaux
    # Pattern pour Tesla
    total_liabilities_pattern = r"Total liabilities\s*\$?\s*([\d,]+)\s*\$?\s*([\d,]+)"
    total_liabilities_match = re.search(total_liabilities_pattern, text)
    
    # Pattern alternatif pour Apple et Microsoft
    alt_total_liabilities_pattern = r"Total liabilities.*?(\$[\d,]+)\s+(\$[\d,]+)"
    alt_total_liabilities_match = re.search(alt_total_liabilities_pattern, text)
    
    if total_liabilities_match:
        metrics['total_liabilities']['2024'] = total_liabilities_match.group(1)
        metrics['total_liabilities']['2023'] = total_liabilities_match.group(2)
        print(f"✅ Passifs totaux extraits avec succès : {metrics['total_liabilities']}")
    elif alt_total_liabilities_match:
        metrics['total_liabilities']['2024'] = alt_total_liabilities_match.group(1)
        metrics['total_liabilities']['2023'] = alt_total_liabilities_match.group(2)
        print(f"✅ Passifs totaux extraits avec succès : {metrics['total_liabilities']}")
    else:
        print("❌ Impossible de trouver les passifs totaux")
    
    # Extraction du flux de trésorerie d'exploitation
    # Pattern pour Tesla
    operating_cash_flow_pattern = r"Cash flows from operating activities.*?Net cash provided by operating activities\s*\$?\s*([\d,]+)\s*\$?\s*([\d,]+)\s*\$?\s*([\d,]+)"
    operating_cash_flow_match = re.search(operating_cash_flow_pattern, text, re.DOTALL)
    
    # Pattern alternatif pour Apple et Microsoft
    alt_operating_cash_flow_pattern = r"Cash generated by operating activities\s*\$?([\d,]+)\s*\$?([\d,]+)\s*\$?([\d,]+)"
    alt_operating_cash_flow_match = re.search(alt_operating_cash_flow_pattern, text)
    
    if operating_cash_flow_match:
        metrics['operating_cash_flow']['2024'] = operating_cash_flow_match.group(1)
        metrics['operating_cash_flow']['2023'] = operating_cash_flow_match.group(2)
        metrics['operating_cash_flow']['2022'] = operating_cash_flow_match.group(3)
        print(f"✅ Flux de trésorerie d'exploitation extraits avec succès : {metrics['operating_cash_flow']}")
    elif alt_operating_cash_flow_match:
        metrics['operating_cash_flow']['2024'] = alt_operating_cash_flow_match.group(1)
        metrics['operating_cash_flow']['2023'] = alt_operating_cash_flow_match.group(2)
        metrics['operating_cash_flow']['2022'] = alt_operating_cash_flow_match.group(3)
        print(f"✅ Flux de trésorerie d'exploitation extraits avec succès : {metrics['operating_cash_flow']}")
    else:
        print("❌ Impossible de trouver les flux de trésorerie d'exploitation")
    
    return metrics

def clean_value(value):
    """Nettoie une valeur financière pour la convertir en nombre."""
    if value is None:
        return 0
    
    # Supprimer le symbole $ et les espaces
    if isinstance(value, str):
        value = value.replace('$', '').replace(' ', '')
        
        # Supprimer les parenthèses (valeurs négatives)
        if '(' in value and ')' in value:
            value = '-' + value.replace('(', '').replace(')', '')
        
        # Supprimer les virgules
        value = value.replace(',', '')
        
        # Convertir en nombre
        try:
            return float(value)
        except ValueError:
            return 0
    
    return float(value)

def calculate_financial_ratios(metrics):
    """Calcule les ratios financiers importants."""
    print("🧮 Calcul des ratios financiers...")
    
    ratios = defaultdict(dict)
    
    for year in ['2024', '2023', '2022']:
        if year in metrics['revenue'] and year in metrics['net_income']:
            # Marge nette
            revenue = clean_value(metrics['revenue'].get(year, 0))
            net_income = clean_value(metrics['net_income'].get(year, 0))
            
            if revenue > 0:
                ratios['net_margin'][year] = (net_income / revenue) * 100
                print(f"✅ Marge nette {year} calculée : {ratios['net_margin'][year]:.2f}%")
        
        if year in metrics['revenue'] and year in metrics['gross_profit']:
            # Marge brute
            revenue = clean_value(metrics['revenue'].get(year, 0))
            gross_profit = clean_value(metrics['gross_profit'].get(year, 0))
            
            if revenue > 0:
                ratios['gross_margin'][year] = (gross_profit / revenue) * 100
                print(f"✅ Marge brute {year} calculée : {ratios['gross_margin'][year]:.2f}%")
    
    # Ratio d'endettement (2024 et 2023 seulement)
    for year in ['2024', '2023']:
        if year in metrics['total_assets'] and year in metrics['total_liabilities']:
            total_assets = clean_value(metrics['total_assets'].get(year, 0))
            total_liabilities = clean_value(metrics['total_liabilities'].get(year, 0))
            
            if total_assets > 0:
                ratios['debt_ratio'][year] = (total_liabilities / total_assets) * 100
                print(f"✅ Ratio d'endettement {year} calculé : {ratios['debt_ratio'][year]:.2f}%")
                
                # Ratio d'autonomie financière (Capitaux propres / Total actif)
                equity = total_assets - total_liabilities
                ratios['financial_autonomy'][year] = (equity / total_assets) * 100
                print(f"✅ Ratio d'autonomie financière {year} calculé : {ratios['financial_autonomy'][year]:.2f}%")
                
                # ROA (Return on Assets) - Rentabilité des actifs
                if year in metrics['net_income']:
                    net_income = clean_value(metrics['net_income'].get(year, 0))
                    ratios['roa'][year] = (net_income / total_assets) * 100
                    print(f"✅ ROA {year} calculé : {ratios['roa'][year]:.2f}%")
                
                # ROE (Return on Equity) - Rentabilité des capitaux propres
                if year in metrics['net_income'] and equity > 0:
                    net_income = clean_value(metrics['net_income'].get(year, 0))
                    ratios['roe'][year] = (net_income / equity) * 100
                    print(f"✅ ROE {year} calculé : {ratios['roe'][year]:.2f}%")
    
    # Calcul des taux de croissance
    for metric in ['revenue', 'net_income', 'gross_profit']:
        if all(year in metrics[metric] for year in ['2022', '2023', '2024']):
            # Taux de croissance annuel composé (CAGR) sur 2 ans
            value_2022 = clean_value(metrics[metric]['2022'])
            value_2024 = clean_value(metrics[metric]['2024'])
            
            if value_2022 > 0:
                cagr = ((value_2024 / value_2022) ** (1/2) - 1) * 100
                # Utiliser une chaîne de caractères comme clé au lieu d'un tuple
                ratios[f'{metric}_cagr']['2022_2024'] = cagr
                print(f"✅ CAGR {metric} (2022-2024) calculé : {cagr:.2f}%")
            
            # Taux de croissance annuels
            value_2023 = clean_value(metrics[metric]['2023'])
            
            if value_2022 > 0:
                growth_2022_2023 = ((value_2023 - value_2022) / value_2022) * 100
                ratios[f'{metric}_growth']['2022_2023'] = growth_2022_2023
                print(f"✅ Croissance {metric} 2022-2023 calculée : {growth_2022_2023:.2f}%")
            
            if value_2023 > 0:
                growth_2023_2024 = ((value_2024 - value_2023) / value_2023) * 100
                ratios[f'{metric}_growth']['2023_2024'] = growth_2023_2024
                print(f"✅ Croissance {metric} 2023-2024 calculée : {growth_2023_2024:.2f}%")
    
    return ratios

def create_visualizations(metrics, ratios, output_dir):
    """Crée des visualisations pour les métriques financières clés."""
    print("📊 Création des visualisations...")
    
    # Création du répertoire de sortie s'il n'existe pas
    os.makedirs(output_dir, exist_ok=True)
    
    # Années disponibles
    years = ['2022', '2023', '2024']
    years_for_assets = ['2023', '2024']  # Seulement 2023 et 2024 pour les actifs et passifs
    
    # Configuration des graphiques
    plt.style.use('ggplot')
    plt.rcParams['figure.figsize'] = (12, 8)
    plt.rcParams['font.size'] = 12
    
    # 1. Graphique des revenus
    plt.figure()
    revenue_values = [clean_value(metrics['revenue'].get(year, 0)) for year in years]
    bars = plt.bar(years, revenue_values, color='#1f77b4')
    plt.title('Revenus de Tesla (en millions $)', fontsize=16)
    plt.xlabel('Année', fontsize=14)
    plt.ylabel('Millions $', fontsize=14)
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    
    # Ajout des valeurs sur les barres
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height + 1000,
                 f'${height:,.0f}', ha='center', va='bottom', fontsize=12)
    
    # Ajout des pourcentages de croissance
    for i in range(1, len(revenue_values)):
        growth = ((revenue_values[i] - revenue_values[i-1]) / revenue_values[i-1]) * 100
        plt.text(i, revenue_values[i] - 5000, f'{growth:.2f}%', 
                 ha='center', va='top', fontsize=12, color='white', fontweight='bold')
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'tesla_revenue.png'), dpi=300)
    plt.close()
    
    # 2. Graphique du bénéfice net
    plt.figure()
    net_income_values = [clean_value(metrics['net_income'].get(year, 0)) for year in years]
    bars = plt.bar(years, net_income_values, color='#2ca02c')
    plt.title('Bénéfice Net de Tesla (en millions $)', fontsize=16)
    plt.xlabel('Année', fontsize=14)
    plt.ylabel('Millions $', fontsize=14)
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    
    # Ajout des valeurs sur les barres
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height + 500,
                 f'${height:,.0f}', ha='center', va='bottom', fontsize=12)
    
    # Ajout des pourcentages de croissance
    for i in range(1, len(net_income_values)):
        growth = ((net_income_values[i] - net_income_values[i-1]) / net_income_values[i-1]) * 100
        plt.text(i, net_income_values[i] - 500, f'{growth:.2f}%', 
                 ha='center', va='top', fontsize=12, color='white', fontweight='bold')
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'tesla_net_income.png'), dpi=300)
    plt.close()
    
    # 3. Graphique des marges
    plt.figure()
    gross_margin_values = [ratios['gross_margin'].get(year, 0) for year in years]
    net_margin_values = [ratios['net_margin'].get(year, 0) for year in years]
    
    plt.plot(years, gross_margin_values, 'o-', linewidth=3, markersize=10, label='Marge Brute')
    plt.plot(years, net_margin_values, 's-', linewidth=3, markersize=10, label='Marge Nette')
    
    plt.title('Évolution des Marges de Tesla (%)', fontsize=16)
    plt.xlabel('Année', fontsize=14)
    plt.ylabel('Pourcentage (%)', fontsize=14)
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.legend(fontsize=12)
    
    # Ajout des valeurs sur les points
    for i, value in enumerate(gross_margin_values):
        plt.text(i, value + 1, f'{value:.2f}%', ha='center', fontsize=12)
    
    for i, value in enumerate(net_margin_values):
        plt.text(i, value - 1, f'{value:.2f}%', ha='center', fontsize=12)
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'tesla_margins.png'), dpi=300)
    plt.close()
    
    # 4. Graphique du ratio d'endettement
    plt.figure()
    debt_ratio_values = [ratios['debt_ratio'].get(year, 0) for year in years_for_assets]
    bars = plt.bar(years_for_assets, debt_ratio_values, color='#d62728')
    plt.title('Ratio d\'Endettement de Tesla (%)', fontsize=16)
    plt.xlabel('Année', fontsize=14)
    plt.ylabel('Pourcentage (%)', fontsize=14)
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    
    # Ajout des valeurs sur les barres
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height + 1,
                 f'{height:.2f}%', ha='center', va='bottom', fontsize=12)
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'tesla_debt_ratio.png'), dpi=300)
    plt.close()
    
    # 5. Graphique comparatif des actifs et passifs
    plt.figure()
    assets_values = [clean_value(metrics['total_assets'].get(year, 0)) for year in years_for_assets]
    liabilities_values = [clean_value(metrics['total_liabilities'].get(year, 0)) for year in years_for_assets]
    
    x = range(len(years_for_assets))
    width = 0.35
    
    plt.bar([i - width/2 for i in x], assets_values, width, label='Actifs Totaux', color='#1f77b4')
    plt.bar([i + width/2 for i in x], liabilities_values, width, label='Passifs Totaux', color='#ff7f0e')
    
    plt.title('Actifs et Passifs de Tesla (en millions $)', fontsize=16)
    plt.xlabel('Année', fontsize=14)
    plt.ylabel('Millions $', fontsize=14)
    plt.xticks(x, years_for_assets)
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.legend(fontsize=12)
    
    # Ajout des valeurs sur les barres
    for i, value in enumerate(assets_values):
        plt.text(i - width/2, value + 2000, f'${value:,.0f}', ha='center', va='bottom', fontsize=12)
    
    for i, value in enumerate(liabilities_values):
        plt.text(i + width/2, value + 2000, f'${value:,.0f}', ha='center', va='bottom', fontsize=12)
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'tesla_assets_liabilities.png'), dpi=300)
    plt.close()
    
    print(f"✅ Visualisations créées et enregistrées dans le répertoire {output_dir}")

def save_results(metrics, ratios, output_dir):
    """Sauvegarde les résultats dans des fichiers CSV et JSON."""
    print("💾 Sauvegarde des résultats...")
    
    # Création du répertoire de sortie s'il n'existe pas
    os.makedirs(output_dir, exist_ok=True)
    
    # Préparation des données pour le CSV
    data = {
        'Métrique': [],
        '2022': [],
        '2023': [],
        '2024': []
    }
    
    # Ajout des métriques financières
    data['Métrique'].append('Revenus (millions $)')
    data['2022'].append(clean_value(metrics['revenue'].get('2022', 0)))
    data['2023'].append(clean_value(metrics['revenue'].get('2023', 0)))
    data['2024'].append(clean_value(metrics['revenue'].get('2024', 0)))
    
    data['Métrique'].append('Bénéfice Net (millions $)')
    data['2022'].append(clean_value(metrics['net_income'].get('2022', 0)))
    data['2023'].append(clean_value(metrics['net_income'].get('2023', 0)))
    data['2024'].append(clean_value(metrics['net_income'].get('2024', 0)))
    
    data['Métrique'].append('Bénéfice Brut (millions $)')
    data['2022'].append(clean_value(metrics['gross_profit'].get('2022', 0)))
    data['2023'].append(clean_value(metrics['gross_profit'].get('2023', 0)))
    data['2024'].append(clean_value(metrics['gross_profit'].get('2024', 0)))
    
    data['Métrique'].append('Actifs Totaux (millions $)')
    data['2022'].append(None)  # Pas de données pour 2022
    data['2023'].append(clean_value(metrics['total_assets'].get('2023', 0)))
    data['2024'].append(clean_value(metrics['total_assets'].get('2024', 0)))
    
    data['Métrique'].append('Passifs Totaux (millions $)')
    data['2022'].append(None)  # Pas de données pour 2022
    data['2023'].append(clean_value(metrics['total_liabilities'].get('2023', 0)))
    data['2024'].append(clean_value(metrics['total_liabilities'].get('2024', 0)))
    
    # Ajout des ratios financiers
    data['Métrique'].append('Marge Brute (%)')
    data['2022'].append(ratios['gross_margin'].get('2022', 0))
    data['2023'].append(ratios['gross_margin'].get('2023', 0))
    data['2024'].append(ratios['gross_margin'].get('2024', 0))
    
    data['Métrique'].append('Marge Nette (%)')
    data['2022'].append(ratios['net_margin'].get('2022', 0))
    data['2023'].append(ratios['net_margin'].get('2023', 0))
    data['2024'].append(ratios['net_margin'].get('2024', 0))
    
    data['Métrique'].append('Ratio d\'Endettement (%)')
    data['2022'].append(None)  # Pas de données pour 2022
    data['2023'].append(ratios['debt_ratio'].get('2023', 0))
    data['2024'].append(ratios['debt_ratio'].get('2024', 0))
    
    # Création du DataFrame et sauvegarde en CSV
    df = pd.DataFrame(data)
    csv_path = os.path.join(output_dir, 'tesla_financial_data.csv')
    df.to_csv(csv_path, index=False)
    
    # Préparation des données pour le JSON
    json_data = {
        'metrics': metrics,
        'ratios': ratios,
        'analysis': {
            'revenue_growth': {
                '2022_to_2023': ((clean_value(metrics['revenue'].get('2023', 0)) - clean_value(metrics['revenue'].get('2022', 0))) / clean_value(metrics['revenue'].get('2022', 0))) * 100,
                '2023_to_2024': ((clean_value(metrics['revenue'].get('2024', 0)) - clean_value(metrics['revenue'].get('2023', 0))) / clean_value(metrics['revenue'].get('2023', 0))) * 100
            },
            'net_income_growth': {
                '2022_to_2023': ((clean_value(metrics['net_income'].get('2023', 0)) - clean_value(metrics['net_income'].get('2022', 0))) / clean_value(metrics['net_income'].get('2022', 0))) * 100,
                '2023_to_2024': ((clean_value(metrics['net_income'].get('2024', 0)) - clean_value(metrics['net_income'].get('2023', 0))) / clean_value(metrics['net_income'].get('2023', 0))) * 100
            },
            'margin_change': {
                'gross_margin': {
                    '2022_to_2023': ratios['gross_margin'].get('2023', 0) - ratios['gross_margin'].get('2022', 0),
                    '2023_to_2024': ratios['gross_margin'].get('2024', 0) - ratios['gross_margin'].get('2023', 0)
                },
                'net_margin': {
                    '2022_to_2023': ratios['net_margin'].get('2023', 0) - ratios['net_margin'].get('2022', 0),
                    '2023_to_2024': ratios['net_margin'].get('2024', 0) - ratios['net_margin'].get('2023', 0)
                }
            },
            'debt_ratio_change': {
                '2023_to_2024': ratios['debt_ratio'].get('2024', 0) - ratios['debt_ratio'].get('2023', 0)
            }
        }
    }
    
    # Sauvegarde en JSON
    json_path = os.path.join(output_dir, 'tesla_financial_data.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=4, ensure_ascii=False)
    
    # Création d'un rapport textuel
    report_path = os.path.join(output_dir, 'tesla_financial_report.txt')
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("RAPPORT D'ANALYSE FINANCIÈRE DE TESLA\n")
        f.write("===================================\n\n")
        
        f.write("MÉTRIQUES FINANCIÈRES CLÉS :\n")
        f.write("---------------------------\n")
        f.write(f"Revenus (en millions $) :\n")
        for year in ['2022', '2023', '2024']:
            f.write(f"  {year}: ${clean_value(metrics['revenue'][year]):,.0f}\n")
        
        f.write(f"\nBénéfice net (en millions $) :\n")
        for year in ['2022', '2023', '2024']:
            f.write(f"  {year}: ${clean_value(metrics['net_income'][year]):,.0f}\n")
        
        f.write(f"\nBénéfice brut (en millions $) :\n")
        for year in ['2022', '2023', '2024']:
            f.write(f"  {year}: ${clean_value(metrics['gross_profit'][year]):,.0f}\n")
        
        f.write(f"\nActifs totaux (en millions $) :\n")
        for year in ['2023', '2024']:
            f.write(f"  {year}: ${clean_value(metrics['total_assets'][year]):,.0f}\n")
        
        f.write(f"\nPassifs totaux (en millions $) :\n")
        for year in ['2023', '2024']:
            f.write(f"  {year}: ${clean_value(metrics['total_liabilities'][year]):,.0f}\n")
        
        f.write("\nRATIOS FINANCIERS :\n")
        f.write("-----------------\n")
        f.write(f"Marge brute (%) :\n")
        for year in ['2022', '2023', '2024']:
            f.write(f"  {year}: {ratios['gross_margin'][year]:.2f}%\n")
        
        f.write(f"\nMarge nette (%) :\n")
        for year in ['2022', '2023', '2024']:
            f.write(f"  {year}: {ratios['net_margin'][year]:.2f}%\n")
        
        f.write(f"\nRatio d'endettement (%) :\n")
        for year in ['2023', '2024']:
            f.write(f"  {year}: {ratios['debt_ratio'][year]:.2f}%\n")
        
        f.write("\nANALYSE DES TENDANCES :\n")
        f.write("---------------------\n")
        
        # Analyse de la croissance des revenus
        revenue_growth_2022_2023 = ((clean_value(metrics['revenue'].get('2023', 0)) - clean_value(metrics['revenue'].get('2022', 0))) / clean_value(metrics['revenue'].get('2022', 0))) * 100
        revenue_growth_2023_2024 = ((clean_value(metrics['revenue'].get('2024', 0)) - clean_value(metrics['revenue'].get('2023', 0))) / clean_value(metrics['revenue'].get('2023', 0))) * 100
        
        f.write(f"Croissance des revenus :\n")
        f.write(f"  2022-2023: {revenue_growth_2022_2023:.2f}%\n")
        f.write(f"  2023-2024: {revenue_growth_2023_2024:.2f}%\n")
        
        if revenue_growth_2023_2024 < revenue_growth_2022_2023:
            f.write(f"  → Ralentissement significatif de la croissance des revenus en 2024.\n")
        else:
            f.write(f"  → Accélération de la croissance des revenus en 2024.\n")
        
        # Analyse de la croissance du bénéfice net
        net_income_growth_2022_2023 = ((clean_value(metrics['net_income'].get('2023', 0)) - clean_value(metrics['net_income'].get('2022', 0))) / clean_value(metrics['net_income'].get('2022', 0))) * 100
        net_income_growth_2023_2024 = ((clean_value(metrics['net_income'].get('2024', 0)) - clean_value(metrics['net_income'].get('2023', 0))) / clean_value(metrics['net_income'].get('2023', 0))) * 100
        
        f.write(f"\nCroissance du bénéfice net :\n")
        f.write(f"  2022-2023: {net_income_growth_2022_2023:.2f}%\n")
        f.write(f"  2023-2024: {net_income_growth_2023_2024:.2f}%\n")
        
        if net_income_growth_2023_2024 < 0:
            f.write(f"  → Baisse significative du bénéfice net en 2024.\n")
        elif net_income_growth_2023_2024 < net_income_growth_2022_2023:
            f.write(f"  → Ralentissement de la croissance du bénéfice net en 2024.\n")
        else:
            f.write(f"  → Accélération de la croissance du bénéfice net en 2024.\n")
        
        # Analyse des marges
        f.write(f"\nÉvolution des marges :\n")
        f.write(f"  Marge brute : {ratios['gross_margin']['2022']:.2f}% (2022) → {ratios['gross_margin']['2023']:.2f}% (2023) → {ratios['gross_margin']['2024']:.2f}% (2024)\n")
        f.write(f"  Marge nette : {ratios['net_margin']['2022']:.2f}% (2022) → {ratios['net_margin']['2023']:.2f}% (2023) → {ratios['net_margin']['2024']:.2f}% (2024)\n")
        
        if ratios['gross_margin']['2024'] < ratios['gross_margin']['2023']:
            f.write(f"  → Détérioration continue de la marge brute.\n")
        else:
            f.write(f"  → Stabilisation ou amélioration de la marge brute en 2024.\n")
        
        if ratios['net_margin']['2024'] < ratios['net_margin']['2023']:
            f.write(f"  → Forte détérioration de la marge nette en 2024.\n")
        else:
            f.write(f"  → Stabilisation ou amélioration de la marge nette en 2024.\n")
        
        # Analyse du ratio d'endettement
        f.write(f"\nÉvolution du ratio d'endettement :\n")
        f.write(f"  {ratios['debt_ratio']['2023']:.2f}% (2023) → {ratios['debt_ratio']['2024']:.2f}% (2024)\n")
        
        if ratios['debt_ratio']['2024'] > ratios['debt_ratio']['2023']:
            f.write(f"  → Augmentation du ratio d'endettement, indiquant un risque financier accru.\n")
        else:
            f.write(f"  → Diminution du ratio d'endettement, indiquant une amélioration de la structure financière.\n")
        
        # Conclusion
        f.write("\nCONCLUSION :\n")
        f.write("-----------\n")
        
        if revenue_growth_2023_2024 > 0 and net_income_growth_2023_2024 < 0:
            f.write("Tesla a connu une année 2024 difficile en termes de rentabilité, malgré une légère croissance des revenus. ")
            f.write("La forte baisse du bénéfice net et des marges suggère des défis opérationnels ou une augmentation des coûts. ")
        elif revenue_growth_2023_2024 <= 0 and net_income_growth_2023_2024 < 0:
            f.write("Tesla a connu une année 2024 très difficile avec une stagnation ou baisse des revenus et une forte diminution de la rentabilité. ")
            f.write("Cette situation pourrait indiquer des problèmes structurels ou une concurrence accrue sur le marché. ")
        elif revenue_growth_2023_2024 > 0 and net_income_growth_2023_2024 > 0:
            f.write("Tesla a maintenu une croissance positive tant au niveau des revenus que du bénéfice net en 2024, ")
            f.write("bien que le rythme de croissance se soit ralenti par rapport à l'année précédente. ")
        
        if ratios['debt_ratio']['2024'] > ratios['debt_ratio']['2023']:
            f.write("L'augmentation du ratio d'endettement indique un risque financier accru, ")
            f.write("bien que le niveau global reste à un niveau raisonnable (inférieur à 40%).")
        else:
            f.write("La structure financière reste solide avec un ratio d'endettement maîtrisé.")
    
    print(f"✅ Résultats sauvegardés dans les fichiers suivants :")
    print(f"   - {csv_path}")
    print(f"   - {json_path}")
    print(f"   - {report_path}")

def process_multiple_reports(file_paths, output_dir):
    """Traite plusieurs rapports financiers et les compare."""
    print("🔄 Traitement de plusieurs rapports financiers...")
    
    all_metrics = {}
    all_ratios = {}
    
    # Traitement de chaque rapport
    for file_path in file_paths:
        company_name = os.path.basename(file_path).split('_')[0].upper()
        print(f"\n📊 Traitement du rapport de {company_name}...")
        
        # Lecture et nettoyage du fichier
        text = read_file(file_path)
        cleaned_text = clean_text(text)
        
        # Extraction des métriques clés
        metrics = extract_key_metrics(cleaned_text)
        
        # Calcul des ratios financiers
        ratios = calculate_financial_ratios(metrics)
        
        # Stockage des résultats
        all_metrics[company_name] = metrics
        all_ratios[company_name] = ratios
        
        # Création des visualisations individuelles
        company_output_dir = os.path.join(output_dir, company_name.lower())
        os.makedirs(company_output_dir, exist_ok=True)
        
        create_visualizations(metrics, ratios, company_output_dir)
        save_results(metrics, ratios, company_output_dir)
    
    # Création de visualisations comparatives
    if len(all_metrics) > 1:
        create_comparative_visualizations(all_metrics, all_ratios, output_dir)
        save_comparative_results(all_metrics, all_ratios, output_dir)
    
    return all_metrics, all_ratios

def create_comparative_visualizations(all_metrics, all_ratios, output_dir):
    """Crée des visualisations comparatives pour plusieurs entreprises."""
    print("📊 Création des visualisations comparatives...")
    
    # Création du répertoire de sortie s'il n'existe pas
    comparative_dir = os.path.join(output_dir, "comparative")
    os.makedirs(comparative_dir, exist_ok=True)
    
    # Années disponibles
    years = ['2022', '2023', '2024']
    
    # Configuration des graphiques
    plt.style.use('ggplot')
    plt.rcParams['figure.figsize'] = (14, 10)
    plt.rcParams['font.size'] = 12
    
    # 1. Comparaison des revenus
    plt.figure()
    
    companies = list(all_metrics.keys())
    x = range(len(years))
    width = 0.8 / len(companies)
    
    for i, company in enumerate(companies):
        revenue_values = [clean_value(all_metrics[company]['revenue'].get(year, 0)) for year in years]
        bars = plt.bar([pos + i * width - (len(companies) - 1) * width / 2 for pos in x], 
                       revenue_values, width, label=company)
    
    plt.title('Comparaison des Revenus (en millions $)', fontsize=16)
    plt.xlabel('Année', fontsize=14)
    plt.ylabel('Millions $', fontsize=14)
    plt.xticks(x, years)
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.legend(fontsize=12)
    
    plt.tight_layout()
    plt.savefig(os.path.join(comparative_dir, 'comparative_revenue.png'), dpi=300)
    plt.close()
    
    # 2. Comparaison des marges nettes
    plt.figure()
    
    for company in companies:
        net_margin_values = [all_ratios[company]['net_margin'].get(year, 0) for year in years]
        plt.plot(years, net_margin_values, 'o-', linewidth=3, markersize=10, label=f"{company} - Marge Nette")
    
    plt.title('Comparaison des Marges Nettes (%)', fontsize=16)
    plt.xlabel('Année', fontsize=14)
    plt.ylabel('Pourcentage (%)', fontsize=14)
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.legend(fontsize=12)
    
    plt.tight_layout()
    plt.savefig(os.path.join(comparative_dir, 'comparative_net_margin.png'), dpi=300)
    plt.close()
    
    # 3. Comparaison des ratios d'endettement
    plt.figure()
    
    years_for_debt = ['2023', '2024']
    x = range(len(years_for_debt))
    width = 0.8 / len(companies)
    
    for i, company in enumerate(companies):
        debt_ratio_values = [all_ratios[company]['debt_ratio'].get(year, 0) for year in years_for_debt]
        bars = plt.bar([pos + i * width - (len(companies) - 1) * width / 2 for pos in x], 
                       debt_ratio_values, width, label=company)
    
    plt.title('Comparaison des Ratios d\'Endettement (%)', fontsize=16)
    plt.xlabel('Année', fontsize=14)
    plt.ylabel('Pourcentage (%)', fontsize=14)
    plt.xticks(x, years_for_debt)
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.legend(fontsize=12)
    
    plt.tight_layout()
    plt.savefig(os.path.join(comparative_dir, 'comparative_debt_ratio.png'), dpi=300)
    plt.close()
    
    print(f"✅ Visualisations comparatives créées et enregistrées dans le répertoire {comparative_dir}")

def save_comparative_results(all_metrics, all_ratios, output_dir):
    """Sauvegarde les résultats comparatifs dans des fichiers CSV et JSON."""
    print("💾 Sauvegarde des résultats comparatifs...")
    
    # Création du répertoire de sortie s'il n'existe pas
    comparative_dir = os.path.join(output_dir, "comparative")
    os.makedirs(comparative_dir, exist_ok=True)
    
    # Préparation des données pour le CSV - Revenus
    revenue_data = {
        'Entreprise': [],
        '2022': [],
        '2023': [],
        '2024': [],
        'Croissance 2022-2023 (%)': [],
        'Croissance 2023-2024 (%)': []
    }
    
    for company in all_metrics.keys():
        revenue_data['Entreprise'].append(company)
        
        revenue_2022 = clean_value(all_metrics[company]['revenue'].get('2022', 0))
        revenue_2023 = clean_value(all_metrics[company]['revenue'].get('2023', 0))
        revenue_2024 = clean_value(all_metrics[company]['revenue'].get('2024', 0))
        
        revenue_data['2022'].append(revenue_2022)
        revenue_data['2023'].append(revenue_2023)
        revenue_data['2024'].append(revenue_2024)
        
        growth_2022_2023 = ((revenue_2023 - revenue_2022) / revenue_2022) * 100 if revenue_2022 > 0 else 0
        growth_2023_2024 = ((revenue_2024 - revenue_2023) / revenue_2023) * 100 if revenue_2023 > 0 else 0
        
        revenue_data['Croissance 2022-2023 (%)'].append(growth_2022_2023)
        revenue_data['Croissance 2023-2024 (%)'].append(growth_2023_2024)
    
    # Création du DataFrame et sauvegarde en CSV - Revenus
    revenue_df = pd.DataFrame(revenue_data)
    revenue_csv_path = os.path.join(comparative_dir, 'comparative_revenue.csv')
    revenue_df.to_csv(revenue_csv_path, index=False)
    
    # Préparation des données pour le CSV - Marges nettes
    net_margin_data = {
        'Entreprise': [],
        '2022 (%)': [],
        '2023 (%)': [],
        '2024 (%)': [],
        'Variation 2022-2023 (points)': [],
        'Variation 2023-2024 (points)': []
    }
    
    for company in all_ratios.keys():
        net_margin_data['Entreprise'].append(company)
        
        margin_2022 = all_ratios[company]['net_margin'].get('2022', 0)
        margin_2023 = all_ratios[company]['net_margin'].get('2023', 0)
        margin_2024 = all_ratios[company]['net_margin'].get('2024', 0)
        
        net_margin_data['2022 (%)'].append(margin_2022)
        net_margin_data['2023 (%)'].append(margin_2023)
        net_margin_data['2024 (%)'].append(margin_2024)
        
        variation_2022_2023 = margin_2023 - margin_2022
        variation_2023_2024 = margin_2024 - margin_2023
        
        net_margin_data['Variation 2022-2023 (points)'].append(variation_2022_2023)
        net_margin_data['Variation 2023-2024 (points)'].append(variation_2023_2024)
    
    # Création du DataFrame et sauvegarde en CSV - Marges nettes
    net_margin_df = pd.DataFrame(net_margin_data)
    net_margin_csv_path = os.path.join(comparative_dir, 'comparative_net_margin.csv')
    net_margin_df.to_csv(net_margin_csv_path, index=False)
    
    # Préparation des données pour le JSON
    comparative_data = {
        'revenue': {},
        'net_margin': {},
        'debt_ratio': {},
        'ranking': {
            'revenue': [],
            'net_margin': [],
            'debt_ratio': []
        }
    }
    
    # Ajout des données de revenus
    for company in all_metrics.keys():
        comparative_data['revenue'][company] = {
            '2022': clean_value(all_metrics[company]['revenue'].get('2022', 0)),
            '2023': clean_value(all_metrics[company]['revenue'].get('2023', 0)),
            '2024': clean_value(all_metrics[company]['revenue'].get('2024', 0))
        }
        
        print(f"{company}: {all_ratios[company]['debt_ratio'].get('2024', 0):.2f}%")
    
    # Ajout des données de marges nettes
    for company in all_ratios.keys():
        comparative_data['net_margin'][company] = {
            '2022': all_ratios[company]['net_margin'].get('2022', 0),
            '2023': all_ratios[company]['net_margin'].get('2023', 0),
            '2024': all_ratios[company]['net_margin'].get('2024', 0),
            'variation_2022_2023': all_ratios[company]['net_margin'].get('2023', 0) - all_ratios[company]['net_margin'].get('2022', 0),
            'variation_2023_2024': all_ratios[company]['net_margin'].get('2024', 0) - all_ratios[company]['net_margin'].get('2023', 0)
        }
    
    # Ajout des données de ratios d'endettement
    for company in all_ratios.keys():
        comparative_data['debt_ratio'][company] = {
            '2023': all_ratios[company]['debt_ratio'].get('2023', 0),
            '2024': all_ratios[company]['debt_ratio'].get('2024', 0),
            'variation_2023_2024': all_ratios[company]['debt_ratio'].get('2024', 0) - all_ratios[company]['debt_ratio'].get('2023', 0)
        }
    
    # Classement des entreprises par revenus 2024
    revenue_ranking = sorted(all_metrics.keys(), key=lambda x: clean_value(all_metrics[x]['revenue'].get('2024', 0)), reverse=True)
    for company in revenue_ranking:
        comparative_data['ranking']['revenue'].append({
            'company': company,
            'value': clean_value(all_metrics[company]['revenue'].get('2024', 0))
        })
    
    # Classement des entreprises par marge nette 2024
    net_margin_ranking = sorted(all_ratios.keys(), key=lambda x: all_ratios[x]['net_margin'].get('2024', 0), reverse=True)
    for company in net_margin_ranking:
        comparative_data['ranking']['net_margin'].append({
            'company': company,
            'value': all_ratios[company]['net_margin'].get('2024', 0)
        })
    
    # Classement des entreprises par ratio d'endettement 2024 (du plus faible au plus élevé)
    debt_ratio_ranking = sorted(all_ratios.keys(), key=lambda x: all_ratios[x]['debt_ratio'].get('2024', 0))
    for company in debt_ratio_ranking:
        comparative_data['ranking']['debt_ratio'].append({
            'company': company,
            'value': all_ratios[company]['debt_ratio'].get('2024', 0)
        })
    
    # Sauvegarde des données en JSON
    json_path = os.path.join(comparative_dir, 'comparative_analysis.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(comparative_data, f, indent=4, ensure_ascii=False)
    
    # Création du rapport textuel
    report_path = os.path.join(comparative_dir, 'comparative_report.txt')
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("RAPPORT D'ANALYSE FINANCIÈRE COMPARATIVE\n")
        f.write("======================================\n\n")
        
        f.write("COMPARAISON DES REVENUS (en millions $) :\n")
        f.write("--------------------------------------\n")
        for company in all_metrics.keys():
            f.write(f"{company} :\n")
            f.write(f"  2022: ${clean_value(all_metrics[company]['revenue'].get('2022', 0)):,.0f}\n")
            f.write(f"  2023: ${clean_value(all_metrics[company]['revenue'].get('2023', 0)):,.0f}\n")
            f.write(f"  2024: ${clean_value(all_metrics[company]['revenue'].get('2024', 0)):,.0f}\n")
            
            growth_2022_2023 = ((clean_value(all_metrics[company]['revenue'].get('2023', 0)) - clean_value(all_metrics[company]['revenue'].get('2022', 0))) / clean_value(all_metrics[company]['revenue'].get('2022', 0))) * 100 if clean_value(all_metrics[company]['revenue'].get('2022', 0)) > 0 else 0
            growth_2023_2024 = ((clean_value(all_metrics[company]['revenue'].get('2024', 0)) - clean_value(all_metrics[company]['revenue'].get('2023', 0))) / clean_value(all_metrics[company]['revenue'].get('2023', 0))) * 100 if clean_value(all_metrics[company]['revenue'].get('2023', 0)) > 0 else 0
            
            f.write(f"  Croissance 2022-2023: {growth_2022_2023:.2f}%\n")
            f.write(f"  Croissance 2023-2024: {growth_2023_2024:.2f}%\n\n")
        
        f.write("\nCOMPARAISON DES MARGES NETTES (%) :\n")
        f.write("----------------------------------\n")
        for company in all_ratios.keys():
            f.write(f"{company} :\n")
            f.write(f"  2022: {all_ratios[company]['net_margin'].get('2022', 0):.2f}%\n")
            f.write(f"  2023: {all_ratios[company]['net_margin'].get('2023', 0):.2f}%\n")
            f.write(f"  2024: {all_ratios[company]['net_margin'].get('2024', 0):.2f}%\n")
            
            variation_2022_2023 = all_ratios[company]['net_margin'].get('2023', 0) - all_ratios[company]['net_margin'].get('2022', 0)
            variation_2023_2024 = all_ratios[company]['net_margin'].get('2024', 0) - all_ratios[company]['net_margin'].get('2023', 0)
            
            f.write(f"  Variation 2022-2023: {variation_2022_2023:.2f} points\n")
            f.write(f"  Variation 2023-2024: {variation_2023_2024:.2f} points\n\n")
        
        f.write("\nCOMPARAISON DES RATIOS D'ENDETTEMENT (%) :\n")
        f.write("----------------------------------------\n")
        for company in all_ratios.keys():
            f.write(f"{company} :\n")
            f.write(f"  2023: {all_ratios[company]['debt_ratio'].get('2023', 0):.2f}%\n")
            f.write(f"  2024: {all_ratios[company]['debt_ratio'].get('2024', 0):.2f}%\n")
            
            variation_2023_2024 = all_ratios[company]['debt_ratio'].get('2024', 0) - all_ratios[company]['debt_ratio'].get('2023', 0)
            
            f.write(f"  Variation 2023-2024: {variation_2023_2024:.2f} points\n\n")
        
        f.write("\nCLASSEMENT DES ENTREPRISES (2024) :\n")
        f.write("--------------------------------\n")
        
        f.write("Par revenus (du plus élevé au plus faible) :\n")
        for i, company_data in enumerate(comparative_data['ranking']['revenue']):
            f.write(f"  {i+1}. {company_data['company']}: ${company_data['value']:,.0f}\n")
        
        f.write("\nPar marge nette (de la plus élevée à la plus faible) :\n")
        for i, company_data in enumerate(comparative_data['ranking']['net_margin']):
            f.write(f"  {i+1}. {company_data['company']}: {company_data['value']:.2f}%\n")
        
        f.write("\nPar ratio d'endettement (du plus faible au plus élevé) :\n")
        for i, company_data in enumerate(comparative_data['ranking']['debt_ratio']):
            f.write(f"  {i+1}. {company_data['company']}: {company_data['value']:.2f}%\n")
        
        f.write("\nCONCLUSION :\n")
        f.write("-----------\n")
        
        # Trouver l'entreprise avec la meilleure croissance des revenus
        best_revenue_growth_company = max(all_metrics.keys(), key=lambda x: ((clean_value(all_metrics[x]['revenue'].get('2024', 0)) - clean_value(all_metrics[x]['revenue'].get('2023', 0))) / clean_value(all_metrics[x]['revenue'].get('2023', 0))) * 100 if clean_value(all_metrics[x]['revenue'].get('2023', 0)) > 0 else 0)
        best_revenue_growth = ((clean_value(all_metrics[best_revenue_growth_company]['revenue'].get('2024', 0)) - clean_value(all_metrics[best_revenue_growth_company]['revenue'].get('2023', 0))) / clean_value(all_metrics[best_revenue_growth_company]['revenue'].get('2023', 0))) * 100 if clean_value(all_metrics[best_revenue_growth_company]['revenue'].get('2023', 0)) > 0 else 0
        
        # Trouver l'entreprise avec la meilleure marge nette
        best_net_margin_company = max(all_ratios.keys(), key=lambda x: all_ratios[x]['net_margin'].get('2024', 0))
        best_net_margin = all_ratios[best_net_margin_company]['net_margin'].get('2024', 0)
        
        # Trouver l'entreprise avec le ratio d'endettement le plus faible
        best_debt_ratio_company = min(all_ratios.keys(), key=lambda x: all_ratios[x]['debt_ratio'].get('2024', 0))
        best_debt_ratio = all_ratios[best_debt_ratio_company]['debt_ratio'].get('2024', 0)
        
        f.write(f"En 2024, {best_revenue_growth_company} a enregistré la meilleure croissance des revenus avec {best_revenue_growth:.2f}%. {best_net_margin_company} a affiché la meilleure marge nette avec {best_net_margin:.2f}%. {best_debt_ratio_company} présente la structure financière la plus solide avec un ratio d'endettement de {best_debt_ratio:.2f}%.\n\n")
        f.write("Cette analyse comparative permet d'identifier les forces et faiblesses relatives de chaque entreprise et de mieux comprendre leur positionnement concurrentiel dans le secteur.")
    
    # Exportation vers Excel
    excel_path = export_to_excel(all_metrics, all_ratios, comparative_dir)
    
    print(f"✅ Résultats comparatifs sauvegardés dans les fichiers suivants :")
    print(f"   - {revenue_csv_path}")
    print(f"   - {net_margin_csv_path}")
    print(f"   - {json_path}")
    print(f"   - {report_path}")
    if excel_path:
        print(f"   - {excel_path}")

def export_to_excel(all_metrics, all_ratios, output_dir):
    """Exporte les données financières comparatives vers un fichier Excel."""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("❌ Les bibliothèques pandas et openpyxl sont requises pour l'exportation Excel.")
        print("   Installez-les avec la commande : pip install pandas openpyxl")
        return None
    
    excel_path = os.path.join(output_dir, "financial_analysis.xlsx")
    
    # Création du classeur Excel
    wb = Workbook()
    
    # Suppression de la feuille par défaut
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Feuille 1: Revenus
    revenue_sheet = wb.create_sheet("Revenus")
    
    # Préparation des données pour les revenus
    revenue_data = {
        'Entreprise': [],
        '2022': [],
        '2023': [],
        '2024': [],
        'Croissance 2022-2023 (%)': [],
        'Croissance 2023-2024 (%)': []
    }
    
    for company in all_metrics.keys():
        revenue_data['Entreprise'].append(company)
        
        revenue_2022 = clean_value(all_metrics[company]['revenue'].get('2022', 0))
        revenue_2023 = clean_value(all_metrics[company]['revenue'].get('2023', 0))
        revenue_2024 = clean_value(all_metrics[company]['revenue'].get('2024', 0))
        
        revenue_data['2022'].append(f"${revenue_2022:,.0f}")
        revenue_data['2023'].append(f"${revenue_2023:,.0f}")
        revenue_data['2024'].append(f"${revenue_2024:,.0f}")
        
        growth_2022_2023 = ((revenue_2023 - revenue_2022) / revenue_2022) * 100 if revenue_2022 > 0 else 0
        growth_2023_2024 = ((revenue_2024 - revenue_2023) / revenue_2023) * 100 if revenue_2023 > 0 else 0
        
        revenue_data['Croissance 2022-2023 (%)'].append(f"{growth_2022_2023:.2f}%")
        revenue_data['Croissance 2023-2024 (%)'].append(f"{growth_2023_2024:.2f}%")
    
    # Conversion en DataFrame
    revenue_df = pd.DataFrame(revenue_data)
    
    # Ajout des données à la feuille
    for r_idx, row in enumerate(dataframe_to_rows(revenue_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = revenue_sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
    
    # Feuille 2: Marges
    margins_sheet = wb.create_sheet("Marges")
    
    # Préparation des données pour les marges
    margins_data = {
        'Entreprise': [],
        'Marge Nette 2022 (%)': [],
        'Marge Nette 2023 (%)': [],
        'Marge Nette 2024 (%)': [],
        'Marge Brute 2022 (%)': [],
        'Marge Brute 2023 (%)': [],
        'Marge Brute 2024 (%)': []
    }
    
    for company in all_ratios.keys():
        margins_data['Entreprise'].append(company)
        
        net_margin_2022 = all_ratios[company]['net_margin'].get('2022', 0)
        net_margin_2023 = all_ratios[company]['net_margin'].get('2023', 0)
        net_margin_2024 = all_ratios[company]['net_margin'].get('2024', 0)
        
        gross_margin_2022 = all_ratios[company]['gross_margin'].get('2022', 0)
        gross_margin_2023 = all_ratios[company]['gross_margin'].get('2023', 0)
        gross_margin_2024 = all_ratios[company]['gross_margin'].get('2024', 0)
        
        margins_data['Marge Nette 2022 (%)'].append(f"{net_margin_2022:.2f}%")
        margins_data['Marge Nette 2023 (%)'].append(f"{net_margin_2023:.2f}%")
        margins_data['Marge Nette 2024 (%)'].append(f"{net_margin_2024:.2f}%")
        
        margins_data['Marge Brute 2022 (%)'].append(f"{gross_margin_2022:.2f}%")
        margins_data['Marge Brute 2023 (%)'].append(f"{gross_margin_2023:.2f}%")
        margins_data['Marge Brute 2024 (%)'].append(f"{gross_margin_2024:.2f}%")
    
    # Conversion en DataFrame
    margins_df = pd.DataFrame(margins_data)
    
    # Ajout des données à la feuille
    for r_idx, row in enumerate(dataframe_to_rows(margins_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = margins_sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
    
    # Feuille 3: Ratios Financiers
    ratios_sheet = wb.create_sheet("Ratios Financiers")
    
    # Préparation des données pour les ratios financiers
    ratios_data = {
        'Entreprise': [],
        'Ratio Endettement 2023 (%)': [],
        'Ratio Endettement 2024 (%)': [],
        'ROA 2023 (%)': [],
        'ROA 2024 (%)': [],
        'ROE 2023 (%)': [],
        'ROE 2024 (%)': []
    }
    
    for company in all_ratios.keys():
        ratios_data['Entreprise'].append(company)
        
        debt_ratio_2023 = all_ratios[company]['debt_ratio'].get('2023', 0)
        debt_ratio_2024 = all_ratios[company]['debt_ratio'].get('2024', 0)
        
        roa_2023 = all_ratios[company]['roa'].get('2023', 0)
        roa_2024 = all_ratios[company]['roa'].get('2024', 0)
        
        roe_2023 = all_ratios[company]['roe'].get('2023', 0)
        roe_2024 = all_ratios[company]['roe'].get('2024', 0)
        
        ratios_data['Ratio Endettement 2023 (%)'].append(f"{debt_ratio_2023:.2f}%")
        ratios_data['Ratio Endettement 2024 (%)'].append(f"{debt_ratio_2024:.2f}%")
        
        ratios_data['ROA 2023 (%)'].append(f"{roa_2023:.2f}%")
        ratios_data['ROA 2024 (%)'].append(f"{roa_2024:.2f}%")
        
        ratios_data['ROE 2023 (%)'].append(f"{roe_2023:.2f}%")
        ratios_data['ROE 2024 (%)'].append(f"{roe_2024:.2f}%")
    
    # Conversion en DataFrame
    ratios_df = pd.DataFrame(ratios_data)
    
    # Ajout des données à la feuille
    for r_idx, row in enumerate(dataframe_to_rows(ratios_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ratios_sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
    
    # Feuille 4: Classement
    ranking_sheet = wb.create_sheet("Classement")
    
    # Préparation des données pour le classement
    ranking_data = {
        'Métrique': [
            'Revenus 2024',
            'Croissance Revenus 2023-2024',
            'Marge Nette 2024',
            'Marge Brute 2024',
            'Ratio Endettement 2024 (le plus bas)',
            'ROA 2024',
            'ROE 2024'
        ],
        '1er': [],
        '2ème': [],
        '3ème': []
    }
    
    # Classement par revenus 2024
    companies_by_revenue = sorted(all_metrics.keys(), 
                                 key=lambda x: clean_value(all_metrics[x]['revenue'].get('2024', 0)), 
                                 reverse=True)
    
    for i, company in enumerate(companies_by_revenue):
        if i == 0:
            ranking_data['1er'].append(f"{company}: ${clean_value(all_metrics[company]['revenue'].get('2024', 0)):,.0f}")
        elif i == 1:
            ranking_data['2ème'].append(f"{company}: ${clean_value(all_metrics[company]['revenue'].get('2024', 0)):,.0f}")
        elif i == 2:
            ranking_data['3ème'].append(f"{company}: ${clean_value(all_metrics[company]['revenue'].get('2024', 0)):,.0f}")
    
    # Classement par croissance des revenus 2023-2024
    def calculate_growth(company):
        revenue_2023 = clean_value(all_metrics[company]['revenue'].get('2023', 0))
        revenue_2024 = clean_value(all_metrics[company]['revenue'].get('2024', 0))
        return ((revenue_2024 - revenue_2023) / revenue_2023) * 100 if revenue_2023 > 0 else 0
    
    companies_by_growth = sorted(all_metrics.keys(), key=calculate_growth, reverse=True)
    
    for i, company in enumerate(companies_by_growth):
        growth = calculate_growth(company)
        if i == 0:
            ranking_data['1er'].append(f"{company}: {growth:.2f}%")
        elif i == 1:
            ranking_data['2ème'].append(f"{company}: {growth:.2f}%")
        elif i == 2:
            ranking_data['3ème'].append(f"{company}: {growth:.2f}%")
    
    # Classement par marge nette 2024
    companies_by_net_margin = sorted(all_ratios.keys(), 
                                    key=lambda x: all_ratios[x]['net_margin'].get('2024', 0), 
                                    reverse=True)
    
    for i, company in enumerate(companies_by_net_margin):
        if i == 0:
            ranking_data['1er'].append(f"{company}: {all_ratios[company]['net_margin'].get('2024', 0):.2f}%")
        elif i == 1:
            ranking_data['2ème'].append(f"{company}: {all_ratios[company]['net_margin'].get('2024', 0):.2f}%")
        elif i == 2:
            ranking_data['3ème'].append(f"{company}: {all_ratios[company]['net_margin'].get('2024', 0):.2f}%")
    
    # Classement par marge brute 2024
    companies_by_gross_margin = sorted(all_ratios.keys(), 
                                      key=lambda x: all_ratios[x]['gross_margin'].get('2024', 0), 
                                      reverse=True)
    
    for i, company in enumerate(companies_by_gross_margin):
        if i == 0:
            ranking_data['1er'].append(f"{company}: {all_ratios[company]['gross_margin'].get('2024', 0):.2f}%")
        elif i == 1:
            ranking_data['2ème'].append(f"{company}: {all_ratios[company]['gross_margin'].get('2024', 0):.2f}%")
        elif i == 2:
            ranking_data['3ème'].append(f"{company}: {all_ratios[company]['gross_margin'].get('2024', 0):.2f}%")
    
    # Classement par ratio d'endettement 2024 (le plus bas est le meilleur)
    companies_by_debt_ratio = sorted(all_ratios.keys(), 
                                    key=lambda x: all_ratios[x]['debt_ratio'].get('2024', 0))
    
    for i, company in enumerate(companies_by_debt_ratio):
        if i == 0:
            ranking_data['1er'].append(f"{company}: {all_ratios[company]['debt_ratio'].get('2024', 0):.2f}%")
        elif i == 1:
            ranking_data['2ème'].append(f"{company}: {all_ratios[company]['debt_ratio'].get('2024', 0):.2f}%")
        elif i == 2:
            ranking_data['3ème'].append(f"{company}: {all_ratios[company]['debt_ratio'].get('2024', 0):.2f}%")
    
    # Classement par ROA 2024
    companies_by_roa = sorted(all_ratios.keys(), 
                             key=lambda x: all_ratios[x]['roa'].get('2024', 0), 
                             reverse=True)
    
    for i, company in enumerate(companies_by_roa):
        if i == 0:
            ranking_data['1er'].append(f"{company}: {all_ratios[company]['roa'].get('2024', 0):.2f}%")
        elif i == 1:
            ranking_data['2ème'].append(f"{company}: {all_ratios[company]['roa'].get('2024', 0):.2f}%")
        elif i == 2:
            ranking_data['3ème'].append(f"{company}: {all_ratios[company]['roa'].get('2024', 0):.2f}%")
    
    # Classement par ROE 2024
    companies_by_roe = sorted(all_ratios.keys(), 
                             key=lambda x: all_ratios[x]['roe'].get('2024', 0), 
                             reverse=True)
    
    for i, company in enumerate(companies_by_roe):
        if i == 0:
            ranking_data['1er'].append(f"{company}: {all_ratios[company]['roe'].get('2024', 0):.2f}%")
        elif i == 1:
            ranking_data['2ème'].append(f"{company}: {all_ratios[company]['roe'].get('2024', 0):.2f}%")
        elif i == 2:
            ranking_data['3ème'].append(f"{company}: {all_ratios[company]['roe'].get('2024', 0):.2f}%")
    
    # Conversion en DataFrame
    ranking_df = pd.DataFrame(ranking_data)
    
    # Ajout des données à la feuille
    for r_idx, row in enumerate(dataframe_to_rows(ranking_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ranking_sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            if c_idx == 1:
                cell.font = Font(bold=True)
    
    # Ajustement de la largeur des colonnes
    for sheet in wb.sheetnames:
        for column in wb[sheet].columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[sheet].column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarde du fichier Excel
    wb.save(excel_path)
    print(f"✅ Données exportées vers Excel : {excel_path}")
    
    return excel_path

def predict_future_performance(metrics, ratios, years_ahead=2):
    """Prédit les performances financières futures en utilisant une régression linéaire."""
    print("🔮 Calcul des performances prédictives...")
    
    try:
        from sklearn.linear_model import LinearRegression
        import numpy as np
    except ImportError:
        print("❌ Les bibliothèques scikit-learn et numpy sont requises pour les prédictions.")
        print("   Installez-les avec la commande : pip install scikit-learn numpy")
        return {}
    
    predictions = {
        'revenue': {},
        'net_income': {},
        'gross_profit': {},
        'net_margin': {},
        'gross_margin': {},
        'confidence': {}
    }
    
    # Années historiques
    historical_years = np.array([2022, 2023, 2024]).reshape(-1, 1)
    future_years = np.array([2025, 2026]).reshape(-1, 1)
    
    # Prédiction des revenus
    if all(year in metrics['revenue'] for year in ['2022', '2023', '2024']):
        revenue_values = np.array([clean_value(metrics['revenue']['2022']), 
                                  clean_value(metrics['revenue']['2023']), 
                                  clean_value(metrics['revenue']['2024'])])
        
        # Créer et entraîner le modèle
        revenue_model = LinearRegression()
        revenue_model.fit(historical_years, revenue_values)
        
        # Calculer le coefficient de détermination (R²)
        r_squared_revenue = revenue_model.score(historical_years, revenue_values)
        predictions['confidence']['revenue'] = r_squared_revenue * 100
        
        # Prédire les revenus futurs
        predicted_revenues = revenue_model.predict(future_years)
        predictions['revenue']['2025'] = predicted_revenues[0]
        predictions['revenue']['2026'] = predicted_revenues[1]
        
        # Calculer le taux de croissance annuel moyen prédit
        predicted_growth_rate = ((predicted_revenues[1] / revenue_values[-1]) ** (1/2) - 1) * 100
        predictions['revenue_growth_rate'] = predicted_growth_rate
        
        print(f"✅ Prédiction revenue pour 2025 : ${predicted_revenues[0]:,.0f}")
        print(f"✅ Prédiction revenue pour 2026 : ${predicted_revenues[1]:,.0f}")
        print(f"✅ Taux de croissance annuel moyen prédit pour revenue : {predicted_growth_rate:.2f}%")
    
    # Prédiction du bénéfice net
    if all(year in metrics['net_income'] for year in ['2022', '2023', '2024']):
        net_income_values = np.array([clean_value(metrics['net_income']['2022']), 
                                     clean_value(metrics['net_income']['2023']), 
                                     clean_value(metrics['net_income']['2024'])])
        
        # Créer et entraîner le modèle
        net_income_model = LinearRegression()
        net_income_model.fit(historical_years, net_income_values)
        
        # Calculer le coefficient de détermination (R²)
        r_squared_net_income = net_income_model.score(historical_years, net_income_values)
        predictions['confidence']['net_income'] = r_squared_net_income * 100
        
        # Prédire le bénéfice net futur
        predicted_net_income = net_income_model.predict(future_years)
        predictions['net_income']['2025'] = predicted_net_income[0]
        predictions['net_income']['2026'] = predicted_net_income[1]
        
        # Calculer le taux de croissance annuel moyen prédit
        predicted_growth_rate = ((predicted_net_income[1] / net_income_values[-1]) ** (1/2) - 1) * 100
        predictions['net_income_growth_rate'] = predicted_growth_rate
        
        print(f"✅ Prédiction net_income pour 2025 : ${predicted_net_income[0]:,.0f}")
        print(f"✅ Prédiction net_income pour 2026 : ${predicted_net_income[1]:,.0f}")
        print(f"✅ Taux de croissance annuel moyen prédit pour net_income : {predicted_growth_rate:.2f}%")
    
    # Prédiction du bénéfice brut
    if all(year in metrics['gross_profit'] for year in ['2022', '2023', '2024']):
        gross_profit_values = np.array([clean_value(metrics['gross_profit']['2022']), 
                                       clean_value(metrics['gross_profit']['2023']), 
                                       clean_value(metrics['gross_profit']['2024'])])
        
        # Créer et entraîner le modèle
        gross_profit_model = LinearRegression()
        gross_profit_model.fit(historical_years, gross_profit_values)
        
        # Calculer le coefficient de détermination (R²)
        r_squared_gross_profit = gross_profit_model.score(historical_years, gross_profit_values)
        predictions['confidence']['gross_profit'] = r_squared_gross_profit * 100
        
        # Prédire le bénéfice brut futur
        predicted_gross_profit = gross_profit_model.predict(future_years)
        predictions['gross_profit']['2025'] = predicted_gross_profit[0]
        predictions['gross_profit']['2026'] = predicted_gross_profit[1]
        
        # Calculer le taux de croissance annuel moyen prédit
        predicted_growth_rate = ((predicted_gross_profit[1] / gross_profit_values[-1]) ** (1/2) - 1) * 100
        predictions['gross_profit_growth_rate'] = predicted_growth_rate
        
        print(f"✅ Prédiction gross_profit pour 2025 : ${predicted_gross_profit[0]:,.0f}")
        print(f"✅ Prédiction gross_profit pour 2026 : ${predicted_gross_profit[1]:,.0f}")
        print(f"✅ Taux de croissance annuel moyen prédit pour gross_profit : {predicted_growth_rate:.2f}%")
    
    # Prédiction des marges
    if 'revenue' in predictions and 'net_income' in predictions and 'gross_profit' in predictions:
        # Marge nette
        for year in ['2025', '2026']:
            if year in predictions['revenue'] and year in predictions['net_income']:
                predictions['net_margin'][year] = (predictions['net_income'][year] / predictions['revenue'][year]) * 100
                print(f"✅ Prédiction net_margin pour {year} : {predictions['net_margin'][year]:.2f}%")
        
        # Marge brute
        for year in ['2025', '2026']:
            if year in predictions['revenue'] and year in predictions['gross_profit']:
                predictions['gross_margin'][year] = (predictions['gross_profit'][year] / predictions['revenue'][year]) * 100
                print(f"✅ Prédiction gross_margin pour {year} : {predictions['gross_margin'][year]:.2f}%")
        
        # Calculer la confiance dans les prédictions de marges
        if all(year in predictions['net_margin'] for year in ['2025', '2026']):
            # Utiliser la moyenne des confiances des métriques utilisées pour calculer la marge
            predictions['confidence']['net_margin'] = (predictions['confidence']['revenue'] + predictions['confidence']['net_income']) / 2
            print(f"✅ Confiance dans la prédiction pour net_margin : {predictions['confidence']['net_margin']:.2f}%")
        
        if all(year in predictions['gross_margin'] for year in ['2025', '2026']):
            predictions['confidence']['gross_margin'] = (predictions['confidence']['revenue'] + predictions['confidence']['gross_profit']) / 2
            print(f"✅ Confiance dans la prédiction pour gross_margin : {predictions['confidence']['gross_margin']:.2f}%")
    
    # Afficher la confiance dans les prédictions
    for metric in ['revenue', 'net_income', 'gross_profit']:
        if metric in predictions['confidence']:
            print(f"✅ Confiance dans la prédiction pour {metric} : {predictions['confidence'][metric]:.2f}%")
    
    return predictions

def create_prediction_visualizations(metrics, ratios, predictions, output_dir):
    """Crée des visualisations pour les prédictions financières."""
    print("📊 Création des visualisations prédictives...")
    
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mtick
    import numpy as np
    import os
    
    # Créer le répertoire de sortie s'il n'existe pas
    os.makedirs(output_dir, exist_ok=True)
    
    # Définir un style professionnel pour les graphiques
    plt.style.use('seaborn-v0_8-darkgrid')
    
    # Couleurs pour les graphiques
    historical_color = '#1f77b4'  # Bleu
    prediction_color = '#ff7f0e'  # Orange
    confidence_color = '#ff7f0e20'  # Orange transparent pour l'intervalle de confiance
    
    # Métriques à visualiser
    metrics_to_visualize = ['revenue', 'net_income', 'gross_profit']
    ratios_to_visualize = ['net_margin', 'gross_margin']
    
    # Années historiques et futures
    historical_years = [2022, 2023, 2024]
    future_years = [2025, 2026]
    all_years = historical_years + future_years
    
    # 1. Visualisation des métriques financières
    for metric in metrics_to_visualize:
        if all(str(year) in metrics[metric] for year in historical_years) and all(str(year) in predictions[metric] for year in future_years):
            plt.figure(figsize=(12, 7))
            
            # Données historiques
            historical_values = [clean_value(metrics[metric][str(year)]) for year in historical_years]
            
            # Données prédites
            predicted_values = [predictions[metric][str(year)] for year in future_years]
            
            # Toutes les valeurs pour l'axe Y
            all_values = historical_values + predicted_values
            
            # Tracer les données historiques
            plt.plot(historical_years, historical_values, 'o-', color=historical_color, linewidth=3, markersize=10, label='Données historiques')
            
            # Tracer les prédictions
            plt.plot(future_years, predicted_values, 'o--', color=prediction_color, linewidth=3, markersize=10, label='Prédictions')
            
            # Ajouter une zone d'intervalle de confiance si disponible
            if metric in predictions['confidence']:
                confidence_level = predictions['confidence'][metric] / 100
                # Calculer un intervalle de confiance simple basé sur le niveau de confiance
                confidence_range = [(1 - confidence_level) * value * 0.5 for value in predicted_values]
                
                plt.fill_between(
                    future_years,
                    [predicted_values[i] - confidence_range[i] for i in range(len(predicted_values))],
                    [predicted_values[i] + confidence_range[i] for i in range(len(predicted_values))],
                    color=confidence_color, alpha=0.5,
                    label=f'Intervalle de confiance ({confidence_level:.2%})'
                )
            
            # Ajouter les valeurs sur les points
            for i, year in enumerate(all_years):
                if i < len(historical_years):
                    value = historical_values[i]
                    plt.annotate(f'${value:,.0f}', 
                                xy=(year, value), 
                                xytext=(0, 10),
                                textcoords='offset points',
                                ha='center', va='bottom',
                                fontsize=9, fontweight='bold')
                else:
                    value = predicted_values[i - len(historical_years)]
                    plt.annotate(f'${value:,.0f}', 
                                xy=(year, value), 
                                xytext=(0, 10),
                                textcoords='offset points',
                                ha='center', va='bottom',
                                fontsize=9, fontweight='bold',
                                color=prediction_color)
            
            # Ajouter le taux de croissance prédit
            if f'{metric}_growth_rate' in predictions:
                growth_rate = predictions[f'{metric}_growth_rate']
                growth_text = f'Taux de croissance annuel moyen prédit: {growth_rate:.2f}%'
                plt.figtext(0.5, 0.01, growth_text, ha='center', fontsize=12, fontweight='bold')
            
            # Formater l'axe Y pour afficher les valeurs en millions/milliards
            def format_y_axis(value, pos):
                if value >= 1e9:
                    return f'${value/1e9:.1f}B'
                elif value >= 1e6:
                    return f'${value/1e6:.1f}M'
                else:
                    return f'${value:.0f}'
            
            plt.gca().yaxis.set_major_formatter(plt.FuncFormatter(format_y_axis))
            
            # Ajouter des titres et des étiquettes
            metric_names = {
                'revenue': 'Revenus',
                'net_income': 'Bénéfice Net',
                'gross_profit': 'Bénéfice Brut'
            }
            
            plt.title(f'Prédiction des {metric_names.get(metric, metric)} (2022-2026)', fontsize=16, fontweight='bold', pad=20)
            plt.xlabel('Année', fontsize=12, labelpad=10)
            plt.ylabel('Valeur ($)', fontsize=12, labelpad=10)
            
            # Personnaliser les ticks de l'axe X
            plt.xticks(all_years, fontsize=10)
            plt.yticks(fontsize=10)
            
            # Ajouter une grille
            plt.grid(True, linestyle='--', alpha=0.7)
            
            # Ajouter une légende
            plt.legend(loc='best', fontsize=10)
            
            # Ajuster les marges
            plt.tight_layout(rect=[0, 0.03, 1, 0.95])
            
            # Sauvegarder le graphique
            output_path = os.path.join(output_dir, f'prediction_{metric}.png')
            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            print(f"✅ Graphique de prédiction pour {metric} sauvegardé dans {output_path}")
            plt.close()
    
    # 2. Visualisation des ratios financiers
    for ratio in ratios_to_visualize:
        if all(str(year) in ratios[ratio] for year in historical_years) and all(str(year) in predictions[ratio] for year in future_years):
            plt.figure(figsize=(12, 7))
            
            # Données historiques
            historical_values = [ratios[ratio][str(year)] for year in historical_years]
            
            # Données prédites
            predicted_values = [predictions[ratio][str(year)] for year in future_years]
            
            # Toutes les valeurs pour l'axe Y
            all_values = historical_values + predicted_values
            
            # Tracer les données historiques
            plt.plot(historical_years, historical_values, 'o-', color=historical_color, linewidth=3, markersize=10, label='Données historiques')
            
            # Tracer les prédictions
            plt.plot(future_years, predicted_values, 'o--', color=prediction_color, linewidth=3, markersize=10, label='Prédictions')
            
            # Ajouter une zone d'intervalle de confiance si disponible
            if ratio in predictions['confidence']:
                confidence_level = predictions['confidence'][ratio] / 100
                # Calculer un intervalle de confiance simple basé sur le niveau de confiance
                confidence_range = [(1 - confidence_level) * value * 0.5 for value in predicted_values]
                
                plt.fill_between(
                    future_years,
                    [predicted_values[i] - confidence_range[i] for i in range(len(predicted_values))],
                    [predicted_values[i] + confidence_range[i] for i in range(len(predicted_values))],
                    color=confidence_color, alpha=0.5,
                    label=f'Intervalle de confiance ({confidence_level:.2%})'
                )
            
            # Ajouter les valeurs sur les points
            for i, year in enumerate(all_years):
                if i < len(historical_years):
                    value = historical_values[i]
                    plt.annotate(f'{value:.2f}%', 
                                xy=(year, value), 
                                xytext=(0, 10),
                                textcoords='offset points',
                                ha='center', va='bottom',
                                fontsize=9, fontweight='bold')
                else:
                    value = predicted_values[i - len(historical_years)]
                    plt.annotate(f'{value:.2f}%', 
                                xy=(year, value), 
                                xytext=(0, 10),
                                textcoords='offset points',
                                ha='center', va='bottom',
                                fontsize=9, fontweight='bold',
                                color=prediction_color)
            
            # Formater l'axe Y pour afficher les pourcentages
            plt.gca().yaxis.set_major_formatter(mtick.PercentFormatter())
            
            # Ajouter des titres et des étiquettes
            ratio_names = {
                'net_margin': 'Marge Nette',
                'gross_margin': 'Marge Brute',
                'debt_ratio': 'Ratio d\'Endettement',
                'roa': 'Rendement des Actifs',
                'roe': 'Rendement des Capitaux Propres'
            }
            
            plt.title(f'Prédiction de {ratio_names.get(ratio, ratio)} (2022-2026)', fontsize=16, fontweight='bold', pad=20)
            plt.xlabel('Année', fontsize=12, labelpad=10)
            plt.ylabel('Pourcentage (%)', fontsize=12, labelpad=10)
            
            # Personnaliser les ticks de l'axe X
            plt.xticks(all_years, fontsize=10)
            plt.yticks(fontsize=10)
            
            # Ajouter une grille
            plt.grid(True, linestyle='--', alpha=0.7)
            
            # Ajouter une légende
            plt.legend(loc='best', fontsize=10)
            
            # Ajuster les marges
            plt.tight_layout()
            
            # Sauvegarder le graphique
            output_path = os.path.join(output_dir, f'prediction_{ratio}.png')
            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            print(f"✅ Graphique de prédiction pour {ratio} sauvegardé dans {output_path}")
            plt.close()
    
    # 3. Créer un graphique combiné pour les métriques principales
    if all(metric in predictions for metric in ['revenue', 'net_income', 'gross_profit']):
        plt.figure(figsize=(14, 8))
        
        # Créer un subplot pour chaque métrique
        fig, axes = plt.subplots(1, 3, figsize=(18, 6))
        
        for i, metric in enumerate(['revenue', 'net_income', 'gross_profit']):
            if all(str(year) in metrics[metric] for year in historical_years) and all(str(year) in predictions[metric] for year in future_years):
                ax = axes[i]
                
                # Données historiques
                historical_values = [clean_value(metrics[metric][str(year)]) for year in historical_years]
                
                # Données prédites
                predicted_values = [predictions[metric][str(year)] for year in future_years]
                
                # Tracer les données historiques
                ax.plot(historical_years, historical_values, 'o-', color=historical_color, linewidth=2, markersize=8, label='Historique')
                
                # Tracer les prédictions
                ax.plot(future_years, predicted_values, 'o--', color=prediction_color, linewidth=2, markersize=8, label='Prédiction')
                
                # Ajouter une zone d'intervalle de confiance si disponible
                if metric in predictions['confidence']:
                    confidence_level = predictions['confidence'][metric] / 100
                    # Calculer un intervalle de confiance simple basé sur le niveau de confiance
                    confidence_range = [(1 - confidence_level) * value * 0.5 for value in predicted_values]
                    
                    ax.fill_between(
                        future_years,
                        [predicted_values[i] - confidence_range[i] for i in range(len(predicted_values))],
                        [predicted_values[i] + confidence_range[i] for i in range(len(predicted_values))],
                        color=confidence_color, alpha=0.5,
                        label=f'Confiance ({confidence_level:.2%})'
                    )
                
                # Formater l'axe Y pour afficher les valeurs en millions/milliards
                def format_y_axis(value, pos):
                    if value >= 1e9:
                        return f'${value/1e9:.1f}B'
                    elif value >= 1e6:
                        return f'${value/1e6:.1f}M'
                    else:
                        return f'${value:.0f}'
                
                ax.yaxis.set_major_formatter(plt.FuncFormatter(format_y_axis))
                
                # Ajouter des titres et des étiquettes
                metric_names = {
                    'revenue': 'Revenus',
                    'net_income': 'Bénéfice Net',
                    'gross_profit': 'Bénéfice Brut'
                }
                
                ax.set_title(f'{metric_names.get(metric, metric)}', fontsize=14, fontweight='bold')
                ax.set_xlabel('Année', fontsize=10)
                
                # Personnaliser les ticks de l'axe X
                ax.set_xticks(all_years)
                ax.set_xticklabels(all_years, fontsize=9, rotation=45)
                
                # Ajouter une grille
                ax.grid(True, linestyle='--', alpha=0.7)
                
                # Ajouter une légende
                ax.legend(loc='best', fontsize=9)
        
        # Ajuster les marges
        plt.tight_layout()
        
        # Sauvegarder le graphique
        output_path = os.path.join(output_dir, 'prediction_combined_metrics.png')
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"✅ Graphique combiné des prédictions sauvegardé dans {output_path}")
        plt.close()
    
    # 4. Créer un graphique combiné pour les ratios
    if all(ratio in predictions for ratio in ['net_margin', 'gross_margin']):
        plt.figure(figsize=(14, 8))
        
        # Créer un subplot pour chaque ratio
        fig, axes = plt.subplots(1, 2, figsize=(14, 6))
        
        for i, ratio in enumerate(['net_margin', 'gross_margin']):
            if all(str(year) in ratios[ratio] for year in historical_years) and all(str(year) in predictions[ratio] for year in future_years):
                ax = axes[i]
                
                # Données historiques
                historical_values = [ratios[ratio][str(year)] for year in historical_years]
                
                # Données prédites
                predicted_values = [predictions[ratio][str(year)] for year in future_years]
                
                # Tracer les données historiques
                ax.plot(historical_years, historical_values, 'o-', color=historical_color, linewidth=2, markersize=8, label='Historique')
                
                # Tracer les prédictions
                ax.plot(future_years, predicted_values, 'o--', color=prediction_color, linewidth=2, markersize=8, label='Prédiction')
                
                # Ajouter une zone d'intervalle de confiance si disponible
                if ratio in predictions['confidence']:
                    confidence_level = predictions['confidence'][ratio] / 100
                    # Calculer un intervalle de confiance simple basé sur le niveau de confiance
                    confidence_range = [(1 - confidence_level) * value * 0.5 for value in predicted_values]
                    
                    ax.fill_between(
                        future_years,
                        [predicted_values[i] - confidence_range[i] for i in range(len(predicted_values))],
                        [predicted_values[i] + confidence_range[i] for i in range(len(predicted_values))],
                        color=confidence_color, alpha=0.5,
                        label=f'Confiance ({confidence_level:.2%})'
                    )
                
                # Formater l'axe Y pour afficher les pourcentages
                ax.yaxis.set_major_formatter(mtick.PercentFormatter())
                
                # Ajouter des titres et des étiquettes
                ratio_names = {
                    'net_margin': 'Marge Nette',
                    'gross_margin': 'Marge Brute'
                }
                
                ax.set_title(f'{ratio_names.get(ratio, ratio)}', fontsize=14, fontweight='bold')
                ax.set_xlabel('Année', fontsize=10)
                
                # Personnaliser les ticks de l'axe X
                ax.set_xticks(all_years)
                ax.set_xticklabels(all_years, fontsize=9, rotation=45)
                
                # Ajouter une grille
                ax.grid(True, linestyle='--', alpha=0.7)
                
                # Ajouter une légende
                ax.legend(loc='best', fontsize=9)
        
        # Ajuster les marges
        plt.tight_layout()
        
        # Sauvegarder le graphique
        output_path = os.path.join(output_dir, 'prediction_combined_ratios.png')
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"✅ Graphique combiné des ratios prédits sauvegardé dans {output_path}")
        plt.close()
    
    print(f"✅ Visualisations prédictives créées avec succès dans {output_dir}")
    return True

def save_prediction_results(predictions, output_dir, ratios):
    """Sauvegarde les résultats des prédictions dans différents formats."""
    print("💾 Sauvegarde des résultats prédictifs...")
    
    import os
    import json
    import csv
    import pandas as pd
    from datetime import datetime
    
    # Créer le répertoire de sortie s'il n'existe pas
    os.makedirs(output_dir, exist_ok=True)
    
    # 1. Sauvegarder les prédictions en JSON
    json_path = os.path.join(output_dir, 'financial_predictions.json')
    with open(json_path, 'w') as f:
        json.dump(predictions, f, indent=4)
    print(f"✅ Prédictions sauvegardées en JSON dans {json_path}")
    
    # 2. Sauvegarder les prédictions en CSV
    # 2.1 Métriques financières
    print(f"\n✅ Extraction des données financières terminée avec succès!")
    print(f"✅ Les résultats ont été sauvegardés dans le répertoire {output_dir}")

def main():
    """Fonction principale qui exécute l'extraction des données financières."""
    print("🚀 Démarrage de l'extraction des données financières...")
    
    # Analyse des arguments de ligne de commande
    parser = argparse.ArgumentParser(description='Extracteur de données financières pour rapports 10-K')
    parser.add_argument('--files', nargs='+', help='Chemins des fichiers à analyser')
    parser.add_argument('--output', default='data/results', help='Répertoire de sortie pour les résultats')
    parser.add_argument('--single', action='store_true', help='Analyser uniquement le fichier Tesla par défaut')
    
    args = parser.parse_args()
    
    # Chemins des fichiers
    output_dir = args.output
    
    # Création du répertoire de sortie s'il n'existe pas
    os.makedirs(output_dir, exist_ok=True)
    
    if args.single or not args.files:
        # Analyse d'un seul fichier (Tesla par défaut)
        input_file = "data/tsla_10k_extracted.txt"
        
        # Lecture et nettoyage du fichier
        text = read_file(input_file)
        cleaned_text = clean_text(text)
        
        # Extraction des métriques clés
        metrics = extract_key_metrics(cleaned_text)
        
        # Calcul des ratios financiers
        ratios = calculate_financial_ratios(metrics)
        
        # Calcul des prédictions
        predictions = predict_future_performance(metrics, ratios)
        
        # Création des visualisations
        create_visualizations(metrics, ratios, output_dir)
        
        # Sauvegarde des résultats
        save_results(metrics, ratios, output_dir)
        
        # Création des visualisations prédictives
        create_prediction_visualizations(metrics, ratios, predictions, output_dir)
        
        # Sauvegarde des résultats prédictifs
        save_prediction_results(predictions, output_dir, ratios)
        
        # Affichage des résultats
        print("\n📊 RÉSULTATS DE L'EXTRACTION DES DONNÉES FINANCIÈRES 📊\n")
        
        print("MÉTRIQUES FINANCIÈRES CLÉS :")
        print("---------------------------")
        print(f"Revenus (en millions $) :")
        for year in ['2022', '2023', '2024']:
            print(f"  {year}: ${clean_value(metrics['revenue'][year]):,.0f}")
        
        print(f"\nBénéfice net (en millions $) :")
        for year in ['2022', '2023', '2024']:
            print(f"  {year}: ${clean_value(metrics['net_income'][year]):,.0f}")
        
        print(f"\nBénéfice brut (en millions $) :")
        for year in ['2022', '2023', '2024']:
            print(f"  {year}: ${clean_value(metrics['gross_profit'][year]):,.0f}")
        
        print(f"\nActifs totaux (en millions $) :")
        for year in ['2023', '2024']:
            if year in metrics['total_assets']:
                print(f"  {year}: ${clean_value(metrics['total_assets'][year]):,.0f}")
            else:
                print(f"  {year}: Non disponible")
        
        print(f"\nPassifs totaux (en millions $) :")
        for year in ['2023', '2024']:
            if year in metrics['total_liabilities']:
                print(f"  {year}: ${clean_value(metrics['total_liabilities'][year]):,.0f}")
            else:
                print(f"  {year}: Non disponible")
        
        print("\nRATIOS FINANCIERS :")
        print("-----------------")
        print(f"Marge brute (%) :")
        for year in ['2022', '2023', '2024']:
            print(f"  {year}: {ratios['gross_margin'][year]:.2f}%")
        
        print(f"\nMarge nette (%) :")
        for year in ['2022', '2023', '2024']:
            print(f"  {year}: {ratios['net_margin'][year]:.2f}%")
        
        print(f"\nRatio d'endettement (%) :")
        for year in ['2023', '2024']:
            if year in ratios['debt_ratio']:
                print(f"  {year}: {ratios['debt_ratio'][year]:.2f}%")
            else:
                print(f"  {year}: Non disponible")
        
        print("\nPRÉDICTIONS FINANCIÈRES :")
        print("----------------------")
        print(f"Revenus prédits (en millions $) :")
        for year in ['2025', '2026']:
            if year in predictions['revenue']:
                print(f"  {year}: ${predictions['revenue'][year]:,.0f}")
        
        print(f"\nBénéfice net prédit (en millions $) :")
        for year in ['2025', '2026']:
            if year in predictions['net_income']:
                print(f"  {year}: ${predictions['net_income'][year]:,.0f}")
        
        print(f"\nMarge nette prédite (%) :")
        for year in ['2025', '2026']:
            if year in predictions['net_margin']:
                print(f"  {year}: {predictions['net_margin'][year]:.2f}%")
        
        print(f"\nTaux de croissance annuel moyen prédit :")
        if 'revenue_growth_rate' in predictions:
            print(f"  Revenus: {predictions['revenue_growth_rate']:.2f}%")
        if 'net_income_growth_rate' in predictions:
            print(f"  Bénéfice net: {predictions['net_income_growth_rate']:.2f}%")
        
    else:
        # Analyse de plusieurs fichiers
        all_metrics, all_ratios = process_multiple_reports(args.files, output_dir)
        
        # Création des visualisations comparatives
        create_comparative_visualizations(all_metrics, all_ratios, os.path.join(output_dir, "comparative"))
        
        # Sauvegarde des résultats comparatifs
        save_comparative_results(all_metrics, all_ratios, os.path.join(output_dir, "comparative"))
    
    print(f"\n✅ Extraction des données financières terminée avec succès!")
    print(f"✅ Les résultats ont été sauvegardés dans le répertoire {output_dir}")

if __name__ == "__main__":
    main() 